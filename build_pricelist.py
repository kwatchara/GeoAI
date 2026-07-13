# -*- coding: utf-8 -*-
"""
build_pricelist.py
------------------
สร้างไฟล์ CSV, XLSX และ HTML (ชื่อเดียวกัน) จากราคาข้อมูลดาวเทียม GISTDA
แหล่งข้อมูล: Gistda_Price_List.pdf (https://www.gistda.or.th/download/Gistda_Price_List.pdf)

โครงสร้าง: ตารางเดียวแบบ normalized, หัวตารางภาษาไทย + อังกฤษ
ข้อมูลถอดความจาก PDF และตรวจสอบซ้ำกับข้อความที่สกัดได้ (validate_against_pdf)
"""
import csv
import io
import os
import html as html_lib

PDF_FILE = "Gistda_Price_List.pdf"
BASENAME = "Gistda_Price_List"

# หัวตาราง (ไทย + อังกฤษ)
COLUMNS = [
    "หมวด (Category)",
    "ระบบ/ดาวเทียม (System/Satellite)",
    "โหมด (Mode)",
    "รายละเอียดภาพ (Resolution)",
    "โพลาไรเซชัน (Polarization)",
    "ราคาในคลัง/ราคาที่ 1 (Archive/Price 1)",
    "ราคาสั่งถ่าย/ราคาที่ 2 (Tasking/Price 2)",
    "หน่วย (Unit)",
    "หมายเหตุ (Notes)",
]

# หมวดหมู่
C_VHR = "รายละเอียดสูงมาก 30–50 ซม. (Very High Resolution Optical)"
C_HR = "รายละเอียดสูง 60 ซม.–2 ม. (High Resolution Optical)"
C_HR2 = "รายละเอียดสูง SPOT / ไทยโชต (High Resolution Optical)"
C_MED = "รายละเอียดปานกลาง >2 ม. (Medium Resolution Optical)"
C_RADAR = "ระบบเรดาร์ (Radar / SAR)"

# แต่ละแถว: (category, system, mode, resolution, polarization, price1, price2, unit, notes)
ROWS = [
    # ---- หน้า 1: Optical รายละเอียดสูงมาก (บาท/ตร.กม.) ----
    (C_VHR, "Pléiades NEO", "", "30 cm.", "", "880", "1,270", "บาท/ตร.กม.",
     "ขั้นต่ำ: คลัง 25 / สั่งถ่าย 100 ตร.กม.; level Primary (PAN, MS, Pansharpened)"),
    (C_VHR, "WorldView-4", "", "30 cm.", "", "920", "1,560", "บาท/ตร.กม.", ""),
    (C_VHR, "SuperView-2", "", "42 cm.", "", "700", "1,100", "บาท/ตร.กม.", ""),
    (C_VHR, "WorldView-1", "", "50 cm.", "", "700", "1,100", "บาท/ตร.กม.", ""),
    (C_VHR, "WorldView-2", "", "50 cm.", "", "700", "1,100", "บาท/ตร.กม.", ""),
    (C_VHR, "WorldView-3", "", "50 cm.", "", "700", "1,100", "บาท/ตร.กม.", ""),
    (C_VHR, "GeoEye-1", "", "50 cm.", "", "700", "1,100", "บาท/ตร.กม.", ""),
    (C_VHR, "Pléiades", "", "50 cm.", "", "490", "830", "บาท/ตร.กม.", ""),
    (C_VHR, "EarthScanner", "", "50 cm.", "", "400", "800", "บาท/ตร.กม.", ""),
    (C_VHR, "SuperView-1", "", "50 cm.", "", "500", "900", "บาท/ตร.กม.", ""),
    (C_VHR, "KOMPSAT-3", "", "50 cm.", "", "400", "700", "บาท/ตร.กม.", ""),
    (C_VHR, "SKYSAT", "", "50 cm.", "", "300", "560", "บาท/ตร.กม.",
     "ขั้นต่ำ 1,250 ตร.กม.; สั่งถ่ายโปรดติดต่อเจ้าหน้าที่; เข้าดูผ่าน API/Explorer"),

    # ---- หน้า 2: Optical รายละเอียดสูง (บาท/ตร.กม.) ----
    (C_HR, "QuickBird", "", "60 cm.", "", "700", "N/A", "บาท/ตร.กม.",
     "ขั้นต่ำ: คลัง 25 / สั่งถ่าย 100 ตร.กม.; level Primary (PAN, MS, Pansharpened)"),
    (C_HR, "GaoFen-7", "", "65 cm.", "", "400", "700", "บาท/ตร.กม.", ""),
    (C_HR, "Jilin", "", "75 cm.", "", "300", "600", "บาท/ตร.กม.", ""),
    (C_HR, "DailyVision", "", "75 cm.", "", "300", "600", "บาท/ตร.กม.", ""),
    (C_HR, "GaoFen-2", "", "80 cm.", "", "300", "400", "บาท/ตร.กม.", ""),
    (C_HR, "IKONOS", "", "1 m.", "", "400", "N/A", "บาท/ตร.กม.", ""),
    (C_HR, "Video Constellation", "", "1 m.", "", "142,500", "285,000", "บาท/30 วินาที",
     "ภาพเคลื่อนไหว/วิดีโอ ≤30 วินาที/ช่วง; ขั้นต่ำ 100 ตร.กม.; แนวถ่าย ≥5 กม."),
    (C_HR, "Night Imaging", "", "1 m.", "", "800", "1,400", "บาท/ตร.กม.",
     "ถ่ายภาพกลางคืน"),

    # ---- หน้า 3: SPOT / ไทยโชต ----
    (C_HR2, "SPOT-6", "", "1.5 m.", "", "190", "230", "บาท/ตร.กม.",
     "ขั้นต่ำ: คลัง 100 / สั่งถ่าย 500 ตร.กม."),
    (C_HR2, "SPOT-7", "", "1.5 m.", "", "190", "230", "บาท/ตร.กม.", ""),
    (C_HR2, "ไทยโชต (THEOS)", "", "2 m.", "", "700", "6,500", "บาท/ภาพ",
     "ราคาปรับ Orthorectification แล้ว 910 บาท/ภาพ"),

    # ---- หน้า 4: Medium (LANDSAT บาท/ภาพ, PLANETSCOPE บาท/ตร.กม./ปี) ----
    (C_MED, "LANDSAT-5", "", "30 m.", "", "150", "N/A", "บาท/ภาพ",
     "Level 1T, 7 Bands; คิดเฉพาะค่าดำเนินการผลิตจากคลัง"),
    (C_MED, "LANDSAT-7", "", "30 m.", "", "150", "N/A", "บาท/ภาพ",
     "Level 1T, 8 Bands"),
    (C_MED, "LANDSAT-8", "", "30 m.", "", "150", "N/A", "บาท/ภาพ",
     "Level 1T, 11 Bands; ให้ สทอภ. ดาวน์โหลด 150 บาท/ภาพ"),
    (C_MED, "LANDSAT-9", "", "30 m.", "", "150", "N/A", "บาท/ภาพ",
     "Level 1T, 11 Bands"),
    (C_MED, "PLANETSCOPE", "Access+Download", "3 m.", "", "180", "240", "บาท/ตร.กม./ปี",
     "คอลัมน์ที่ 2 = การติดตาม (Monitoring); สัญญา 1 ปี; ขั้นต่ำ 100 ตร.กม.; "
     "เข้าดู/ดาวน์โหลดผ่าน Planet Explorer, Planet API, Desktop GIS"),

    # ---- หน้า 5: RADARSAT-2 (C band) — Price1=Single Look complex, Price2=Path Image ----
    (C_RADAR, "RADARSAT-2 (C band)", "Standard", "25 m.", "", "57,600", "57,600", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "Spotlight A", "1 m.", "", "134,400", "134,400", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "Utra-Fine", "3 m.", "", "86,400", "86,400", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "Wide Utra-Fine", "3 m.", "", "124,800", "124,800", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "Multi-Look Fine", "8 m.", "", "67,200", "67,200", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "Wide Multi-Look Fine", "8 m.", "", "120,000", "120,000", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "Fine", "8 m.", "", "57,600", "57,600", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "Wide", "30 m.", "", "57,600", "57,600", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "ScanSAR Narrow", "50 m.", "", "N/A", "57,600", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "ScanSAR Wide", "100 m.", "", "N/A", "57,600", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "Extended High, Low", "25 m.", "", "57,600", "57,600", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "Fine Quad-Pol", "8 m.", "", "86,400", "N/A", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),
    (C_RADAR, "RADARSAT-2 (C band)", "Wide Fine Quad-Pol", "8 m.", "", "124,800", "N/A", "บาท/ภาพ",
     "ราคา: Single Look complex | Path Image"),

    # ---- หน้า 6: TerraSAR-X (X band) — Archive/Tasking ----
    (C_RADAR, "TerraSAR-X (X band)", "Staring Spotlight (ST)", "0.25 m.", "", "162,630", "325,260", "บาท/ภาพ", ""),
    (C_RADAR, "TerraSAR-X (X band)", "High Res Spotlight (HS)", "1 m.", "", "139,230", "278,460", "บาท/ภาพ", ""),
    (C_RADAR, "TerraSAR-X (X band)", "Spotlight", "2 m.", "", "99,450", "198,900", "บาท/ภาพ", ""),
    (C_RADAR, "TerraSAR-X (X band)", "StripMap", "3 m.", "", "69,030", "138,060", "บาท/ภาพ", ""),
    (C_RADAR, "TerraSAR-X (X band)", "ScanSAR", "18.5 m.", "", "40,950", "81,900", "บาท/ภาพ", ""),
    (C_RADAR, "TerraSAR-X (X band)", "Wide ScanSAR", "40 m.", "", "40,950", "81,900", "บาท/ภาพ", ""),

    # ---- หน้า 6: COSMO SkyMed (X band) — Price2 = New Acquisition ----
    (C_RADAR, "COSMO SkyMed (X band)", "Spotlight-2", "1x1 m.", "HH, VV", "N/A", "180,000", "บาท/ภาพ",
     "New Acquisition"),
    (C_RADAR, "COSMO SkyMed (X band)", "StripMap Himage", "3 x 3 – 5 x 5 m.", "HH, HV, VH, VV", "N/A", "93,000", "บาท/ภาพ",
     "New Acquisition"),
    (C_RADAR, "COSMO SkyMed (X band)", "StripMap PingPong", "10 x 12 – 20 x 20 m.",
     "HH,VV / HH,HV / VV,VH (2 ช่องสัญญาณ polarimetric)", "N/A", "68,000", "บาท/ภาพ", "New Acquisition"),
    (C_RADAR, "COSMO SkyMed (X band)", "ScanSAR Wide", "14 x 22 – 30 x 30 m.", "HH, HV, VH, VV", "N/A", "78,000", "บาท/ภาพ",
     "New Acquisition"),
    (C_RADAR, "COSMO SkyMed (X band)", "ScanSAR Huge", "14 x 38 – 100 x 100 m.", "HH, HV, VH, VV", "N/A", "78,000", "บาท/ภาพ",
     "New Acquisition"),

    # ---- หน้า 7: GaoFen-3 (C band) — Archive/Tasking ----
    (C_RADAR, "GaoFen-3 (C band)", "Spotlight (SL)", "1 m.", "HH, VV", "116,400", "180,500", "บาท/ภาพ", ""),
    (C_RADAR, "GaoFen-3 (C band)", "Ultra-fine Stripmap (UFS)", "3 m.", "HH, VV", "68,900", "118,800", "บาท/ภาพ", ""),
    (C_RADAR, "GaoFen-3 (C band)", "Fine Stripmap (FSI)", "5 m.", "HH, VV", "64,200", "95,000", "บาท/ภาพ", ""),
    (C_RADAR, "GaoFen-3 (C band)", "Wide Fine Stripmap (FSII)", "10 m.", "HH, HV / VV, VH", "64,200", "90,300", "บาท/ภาพ", ""),
    (C_RADAR, "GaoFen-3 (C band)", "Standard Stripmap (SS)", "25 m.", "HH, HV / VV, VH", "54,700", "85,500", "บาท/ภาพ", ""),
    (C_RADAR, "GaoFen-3 (C band)", "Narrow ScanSAR (NSC)", "50 m.", "HH, HV / VV, VH", "32,100", "42,800", "บาท/ภาพ", ""),
    (C_RADAR, "GaoFen-3 (C band)", "Wide ScanSAR (WSC)", "100 m.", "HH, HV / VV, VH", "32,100", "45,800", "บาท/ภาพ", ""),
    (C_RADAR, "GaoFen-3 (C band)", "Quad-pol Stripmap (QPSI)", "8 m.", "HH, HV / VV, VH", "71,300", "137,800", "บาท/ภาพ", ""),
    (C_RADAR, "GaoFen-3 (C band)", "Wide Quad-pol Stripmap (QPSII)", "25 m.", "HH, HV / VV, VH", "71,300", "137,800", "บาท/ภาพ", ""),
    (C_RADAR, "GaoFen-3 (C band)", "Wave (WAV)", "10 m.", "HH, HV / VV, VH", "10,700", "14,300", "บาท/ภาพ", ""),
    (C_RADAR, "GaoFen-3 (C band)", "Global Observation (GLO)", "500 m.", "HH, HV / VV, VH", "10,700", "14,300", "บาท/ภาพ", ""),
    (C_RADAR, "GaoFen-3 (C band)", "Extended Incidence Angle (EXT)", "25 m.", "HH, HV / VV, VH", "42,800", "57,000", "บาท/ภาพ", ""),
]

FOOTNOTE = ("**ราคาดังกล่าวยังไม่รวมภาษีมูลค่าเพิ่ม | "
            "สอบถามเพิ่มเติม: สทอภ. (GISTDA) ฝ่ายพัฒนาธุรกิจและการบริการ "
            "โทร 0 2143 9593, 0 2141 4564-66,69 | usd@gistda.or.th")


def validate_against_pdf():
    """ตรวจสอบว่าเลขราคาและชื่อระบบทุกค่าปรากฏจริงในข้อความที่สกัดจาก PDF."""
    try:
        import pdfplumber
    except ImportError:
        print("  [ข้าม validation] ไม่พบ pdfplumber")
        return
    if not os.path.exists(PDF_FILE):
        print(f"  [ข้าม validation] ไม่พบ {PDF_FILE}")
        return
    with pdfplumber.open(PDF_FILE) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    missing = []
    for r in ROWS:
        _, system, mode, _, _, p1, p2, _, _ = r
        checks = [system.split(" (")[0].split(" (")[0]]  # ชื่อฐาน (อังกฤษ)
        if mode:
            checks.append(mode.split(" (")[0])
        for price in (p1, p2):
            if price and price != "N/A":
                checks.append(price)
        for token in checks:
            # ข้ามโทเคนภาษาไทย (การสกัด PDF มีปัญหาช่องว่างในภาษาไทย)
            if any("฀" <= ch <= "๿" for ch in token):
                continue
            if token and token not in text:
                missing.append(f"{system} / {mode or '-'} -> '{token}'")
    if missing:
        print(f"  [เตือน] {len(missing)} โทเคนไม่พบใน PDF:")
        for m in missing:
            print("    -", m)
    else:
        print("  [OK] ทุกชื่อระบบ/โหมด/ราคา (อังกฤษ+ตัวเลข) พบใน PDF ครบ")


def write_csv():
    path = f"{BASENAME}.csv"
    with io.open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(COLUMNS)
        w.writerows(ROWS)
        w.writerow([])
        w.writerow([FOOTNOTE])
    print(f"  เขียน {path} ({len(ROWS)} แถว)")


def write_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    path = f"{BASENAME}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "GISTDA Price List"

    title = "ราคาข้อมูลจากดาวเทียม GISTDA (Gistda Price List)"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.cell(1, 1).font = Font(bold=True, size=14, color="1F4E79")
    ws.cell(1, 1).alignment = Alignment(horizontal="left")

    header_row = 2
    ws.append(COLUMNS)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(header_row, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    alt_fill = PatternFill("solid", fgColor="F2F6FC")
    for i, r in enumerate(ROWS):
        ws.append(list(r))
        row_idx = header_row + 1 + i
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row_idx, c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (1, 9)))
            if i % 2 == 1:
                cell.fill = alt_fill
        # จัดชิดขวาสำหรับคอลัมน์ราคา (6, 7)
        for c in (6, 7):
            ws.cell(row_idx, c).alignment = Alignment(horizontal="right", vertical="center")

    foot_row = header_row + 1 + len(ROWS) + 1
    ws.cell(foot_row, 1, FOOTNOTE)
    ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=len(COLUMNS))
    ws.cell(foot_row, 1).font = Font(italic=True, size=9, color="555555")
    ws.cell(foot_row, 1).alignment = Alignment(wrap_text=True)

    widths = [34, 22, 26, 20, 34, 18, 18, 16, 44]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    wb.save(path)
    print(f"  เขียน {path}")


def write_html():
    path = f"{BASENAME}.html"
    esc = html_lib.escape

    # จัดกลุ่มตามหมวด โดยคงลำดับ
    order = []
    groups = {}
    for r in ROWS:
        cat = r[0]
        if cat not in groups:
            groups[cat] = []
            order.append(cat)
        groups[cat].append(r)

    parts = []
    parts.append("""<meta charset="utf-8">
<style>
  :root { --brand:#1f4e79; --brand2:#2e6da4; --line:#e2e8f0; --alt:#f5f8fc; --ink:#1a202c; --mut:#64748b; }
  * { box-sizing: border-box; }
  body { font-family: 'Segoe UI','Sarabun',Tahoma,sans-serif; color:var(--ink); margin:0; padding:24px; background:#fff; line-height:1.45; }
  h1 { color:var(--brand); font-size:1.55rem; margin:0 0 4px; }
  .sub { color:var(--mut); margin:0 0 20px; font-size:.92rem; }
  h2 { color:var(--brand); font-size:1.1rem; margin:28px 0 8px; padding-bottom:4px; border-bottom:2px solid var(--brand); }
  .tbl-wrap { overflow-x:auto; }
  table { border-collapse:collapse; width:100%; font-size:.9rem; margin-bottom:8px; }
  th,td { border:1px solid var(--line); padding:7px 10px; text-align:left; vertical-align:top; }
  thead th { background:var(--brand); color:#fff; position:sticky; top:0; font-weight:600; }
  tbody tr:nth-child(even) { background:var(--alt); }
  td.num { text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap; }
  .na { color:var(--mut); }
  .note { color:var(--mut); font-size:.82rem; }
  footer { margin-top:28px; padding-top:12px; border-top:1px solid var(--line); color:var(--mut); font-size:.82rem; }
  @media print { thead th { position:static; } body { padding:0; } }
</style>
""")
    parts.append('<h1>ราคาข้อมูลจากดาวเทียม GISTDA</h1>')
    parts.append('<p class="sub">Gistda Price List — ที่มา: Gistda_Price_List.pdf | '
                 f'จำนวน {len(ROWS)} รายการ</p>')

    head_cells = "".join(f"<th>{esc(c)}</th>" for c in COLUMNS[1:])  # ตัดคอลัมน์หมวดออก (ใช้เป็นหัวข้อ)
    for cat in order:
        parts.append(f"<h2>{esc(cat)}</h2>")
        parts.append('<div class="tbl-wrap"><table>')
        parts.append(f"<thead><tr>{head_cells}</tr></thead><tbody>")
        for r in groups[cat]:
            tds = []
            for idx, val in enumerate(r[1:], start=1):  # ข้าม category
                v = esc(val) if val else ""
                if idx in (5, 6):  # ราคา (index ใน r[1:]: 5,6)
                    cls = "num"
                    if val == "N/A":
                        tds.append(f'<td class="num na">N/A</td>')
                        continue
                    tds.append(f'<td class="{cls}">{v}</td>')
                elif idx == 8:  # notes
                    tds.append(f'<td class="note">{v}</td>')
                else:
                    tds.append(f"<td>{v}</td>")
            parts.append("<tr>" + "".join(tds) + "</tr>")
        parts.append("</tbody></table></div>")

    parts.append(f'<footer>{esc(FOOTNOTE)}</footer>')

    with io.open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(parts))
    print(f"  เขียน {path}")


def main():
    print("1) ตรวจสอบข้อมูลกับ PDF:")
    validate_against_pdf()
    print("2) สร้างไฟล์ output:")
    write_csv()
    write_xlsx()
    write_html()
    print("เสร็จสิ้น.")


if __name__ == "__main__":
    main()
